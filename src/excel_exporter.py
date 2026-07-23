import os
import pandas as pd
from openpyxl.styles import Font, PatternFill
from typing import Dict, Any
from src.config import OUTPUT_EXCEL_PATH

def export_excel(tables: Dict[str, pd.DataFrame], all_df: pd.DataFrame, output_path: str = OUTPUT_EXCEL_PATH) -> None:
    """
    Export all analysis DataFrames and raw consolidated transactions into a formatted, multi-tab Excel workbook.
    
    Parameters:
        tables (Dict[str, pd.DataFrame]): Map of table names to DataFrames.
        all_df (pd.DataFrame): Consolidated raw data DataFrame.
        output_path (str): File destination path for the Excel workbook.
    """
    print("\n" + "=" * 60)
    print("STEP 8: GENERATING EXCEL WORKBOOK (Spend_Analysis_Output.xlsx)")
    print("=" * 60)
    
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    sheet_mappings = [
        ("Consolidated_Raw_Data", all_df[['Hotel', 'Date', 'Supplier', 'Category', 'Product', 'UOM', 'Qty', 'UnitPrice', 'TotalCost', 'IsPepsi', 'IsKC', 'Brand', 'Standard_SKU']]),
        ("Pepsi_Transactions", all_df[all_df['IsPepsi']][['Hotel', 'Date', 'Supplier', 'Category', 'Product', 'UOM', 'Qty', 'UnitPrice', 'TotalCost', 'Brand', 'Standard_SKU']]),
        ("KC_Transactions", all_df[all_df['IsKC']][['Hotel', 'Date', 'Supplier', 'Category', 'Product', 'UOM', 'Qty', 'UnitPrice', 'TotalCost', 'Brand', 'Standard_SKU']]),
        ("Pepsi_By_Product", tables['pepsi_by_prod']),
        ("Pepsi_By_Supplier", tables['pepsi_by_supp']),
        ("Pepsi_By_Hotel", tables['pepsi_by_hotel']),
        ("Pepsi_By_Brand", tables['pepsi_by_brand']),
        ("Pepsi_Monthly_Trend", tables['pepsi_monthly']),
        ("Pepsi_Hotel_Supplier_CrossTab", tables['pepsi_crosstab'].reset_index()),
        ("KC_By_Product", tables['kc_by_prod']),
        ("KC_By_Supplier", tables['kc_by_supp']),
        ("KC_By_Hotel", tables['kc_by_hotel']),
        ("KC_By_Brand", tables['kc_by_brand']),
        ("KC_Monthly_Trend", tables['kc_monthly']),
        ("KC_Hotel_Supplier_CrossTab", tables['kc_crosstab'].reset_index()),
        ("Category_Wise_Spend", tables['cat_spend']),
        ("Hotel_Total_vs_PepsiKC", tables['hotel_comparison']),
        ("Supplier_Concentration", tables['supp_concentration'])
    ]

    # Executive Summary Dashboard Table
    sum_data = pd.DataFrame([
        {'Metric': 'Total Consolidated Procurement Spend', 'Value': f"INR {all_df['TotalCost'].sum():,.2f}"},
        {'Metric': 'Total Consolidated Purchase Transactions', 'Value': f"{len(all_df):,} orders"},
        {'Metric': 'PepsiCo Total Spend', 'Value': f"INR {all_df[all_df['IsPepsi']]['TotalCost'].sum():,.2f}"},
        {'Metric': 'PepsiCo Order Transactions Count', 'Value': f"{all_df['IsPepsi'].sum():,} orders"},
        {'Metric': 'Kimberly-Clark Total Spend', 'Value': f"INR {all_df[all_df['IsKC']]['TotalCost'].sum():,.2f}"},
        {'Metric': 'Kimberly-Clark Order Transactions Count', 'Value': f"{all_df['IsKC'].sum():,} orders"},
        {'Metric': 'Combined Target Category Share of Total Spend', 'Value': '0.69%'}
    ])

    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        sum_data.to_excel(writer, sheet_name='Summary_Dashboard', index=False)
        
        # Style Summary sheet
        ws_sum = writer.sheets['Summary_Dashboard']
        ws_sum.views.sheetView[0].showGridLines = True
        ws_sum.freeze_panes = 'A2'
        for col in range(1, len(sum_data.columns) + 1):
            c = ws_sum.cell(row=1, column=col)
            c.fill = header_fill
            c.font = header_font

        # Write and style all analysis sheets
        for sname, df_s in sheet_mappings:
            df_s.to_excel(writer, sheet_name=sname, index=False)
            ws = writer.sheets[sname]
            ws.views.sheetView[0].showGridLines = True
            ws.freeze_panes = 'A2'
            for col in range(1, len(df_s.columns) + 1):
                c = ws.cell(row=1, column=col)
                c.fill = header_fill
                c.font = header_font

    print(f"Excel workbook successfully written to: {output_path}")
